
import openpyxl as xl

def write_to_excel(full_file_path, search_column, search_value, set_column, set_value):
    wb = None
    try:
        wb = xl.load_workbook(full_file_path, read_only=False)
    except FileNotFoundError:
        print("File not found.")

    ws = wb.worksheets[0]
    max_column = ws.max_column
    max_row = ws.max_row
    headers = {}
    for i in range(1, max_column + 1):
        cell = ws.cell(row=1, column=i)
        headers[cell.value] = i
    
    for row in range(1, max_row + 1):
        current_row = ws.cell(row=row, column=headers[search_column])
        if (current_row.value == search_value):
            ws.cell(row=row, column=headers[set_column]).value = set_value
            wb.save(full_file_path)
            print("Saved.")
            return True
    
    return False

#b = write_to_excel('ID',1175,'Automation Status', 'Completed')
#print ("Success = " + str(b))


def writeToExcel(**kwargs):
    """
    kwargs = {
        full_file_path,
        search_column,
        search_value,
        set_column,
        set_value
    }
    """
    wb = None
    try:
        wb = xl.load_workbook(kwargs["full_file_path"], read_only=False)
    except FileNotFoundError:
        print("File not found.")

    ws = wb.worksheets[0]
    max_column = ws.max_column
    max_row = ws.max_row
    headers = {}
    for i in range(1, max_column + 1):
        cell = ws.cell(row=1, column=i)
        headers[cell.value] = i
    
    for row in range(1, max_row + 1):
        current_row = ws.cell(row=row, column=headers[kwargs["search_column"]])
        if (current_row.value == kwargs["search_value"]):
            ws.cell(row=row, column=headers[kwargs["set_column"]]).value = kwargs["set_value"]
            wb.save(kwargs["full_file_path"])
            print("Saved.")
            return True
    
    return False



def addRowToExcel(**kwargs):
    """
    kwargs = {
        full_file_path,
        sheet_name
        row_data {
            "header1": "value1", 
            ... ,
            "header100": "value100"
        }
    }
    """
    wb = None
    try:
        wb = xl.load_workbook(kwargs["full_file_path"], read_only=False)
    except FileNotFoundError:
        print("File not found.")
        exit()

    ws = None
    if(kwargs["sheet_name"] in wb.get_sheet_names()):
        ws = ws = wb.get_sheet_by_name(kwargs["sheet_name"])
    else:
        ws = wb.worksheets[0]

    max_column = ws.max_column
    max_row = ws.max_row
    headers = {}
    for i in range(1, max_column + 1):
        cell = ws.cell(row=1, column=i)
        headers[cell.value] = i

    for key in kwargs["my_data"]:
        val = kwargs["my_data"][key]
        if(key in headers):
            ws.cell(row = max_row + 1, column = headers[key]).value = val
            print(f"Column: {key}, Value: {val}")
        else:
            msg_header_not_found = f"Error: the header {key} , not in Excel file"
            print(msg_header_not_found)
    
    try:
        wb.save(kwargs["full_file_path"])
        print("Saved.")
        return True
    except:
        print("Error saving Excel file.")
        return False

# addRowToExcel(
#     full_file_path = R"C:\Users\meirg\OneDrive - Nice Systems Ltd\LMS\HTML CSS Javascript\csod_scrape_emails.xlsx",
#     sheet_name="Sheet1",
#     my_data= {
#         # "Action": "My action lala",
#         "Action Type": "my action type lala",
#         "Action Description1": "My action description",
#         "Action Description": "showing yuval how it works"
#     }
# )    